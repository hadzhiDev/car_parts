from django.contrib import admin
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rangefilter.filters import DateTimeRangeFilter, DateRangeFilterBuilder

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import Sale, SaleItem, Client, Payment
from apps.main.utils import convert_from_usd


def parse_admin_date(value):
    """
    Supports:
    - YYYY-MM-DD (ISO)
    - DD.MM.YYYY (admin UI)
    """
    if not value:
        return None

    date = parse_date(value)
    if date:
        return date

    try:
        return datetime.strptime(value, '%d.%m.%Y').date()
    except ValueError:
        return None



@admin.action(description='Экспорт выбранных товаров продажи в Excel')
def export_sale_items_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары продажи"

    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    left = Alignment(horizontal='left')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ================= DATE RANGE =================
    start = request.GET.get('sale__sale_date__range__gte')
    end = request.GET.get('sale__sale_date__range__lte')

    start_date = parse_admin_date(start)
    end_date = parse_admin_date(end)

    date_range_text = "Период: "
    if start_date and end_date:
        date_range_text += f"{start_date} — {end_date}"
    else:
        date_range_text += "за всё время"

    # ================= TOTAL AMOUNT =================
    total_amount = sum(item.total_cost for item in queryset)

    # ================= TOP INFO =================
    ws.merge_cells('D1:H1')
    ws['D1'] = date_range_text
    ws['D1'].font = bold
    ws['D1'].alignment = left

    ws.merge_cells('D2:H2')
    ws['D2'] = f"Общая сумма продаж: {total_amount}"
    ws['D2'].font = bold
    ws['D2'].alignment = left

    # ================= HEADERS =================
    headers = [
        "Продажа ID",
        "Дата продажи",
        "Клиент",
        "Товар",
        "Артикул",
        "Количество",
        "Цена продажи",
        "Сумма",
    ]

    ws.append([])  # empty row
    ws.append(headers)

    header_row = ws.max_row

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border

    # ================= DATA =================
    for item in queryset.select_related('sale', 'sale__client', 'product'):
        ws.append([
            item.sale.id,
            item.sale.sale_date.strftime('%Y-%m-%d %H:%M'),
            item.sale.client.full_name,
            item.product.name,
            getattr(item.product, 'article_number', ''),
            item.quantity,
            float(item.sale_price),
            float(item.total_cost),
        ])

    # ================= STYLING BODY =================
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = left if cell.column in [3, 4] else center

    # ================= AUTO WIDTH =================
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # ================= RESPONSE =================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sale_items.xlsx"'
    wb.save(response)

    return response




class SaleItemInline(admin.TabularInline):
    model = SaleItem
    extra = 1
    autocomplete_fields = ('product',)
    readonly_fields = ('row_total',)
    fields = (
        'product',
        'quantity',
        'sale_price',
        'row_total',
    )
    def row_total(self, obj):
        print('obj.total_cost: ', obj.total_cost)
        return "—"
    row_total.short_description = "Итого"

 
@admin.register(Sale)
class SaleAdmin(admin.ModelAdmin):
    list_display = ('id', 'sale_date', 'client__full_name', 'total_amount')
    list_display_links = ('id', 'sale_date')
    list_filter = ('sale_date', 'client__full_name')
    search_fields = ('id', 'client__full_name')
    inlines = [SaleItemInline]
    autocomplete_fields = ('client',)

    @admin.display(description='Общая сумма')
    def total_amount(self, obj):
        return convert_from_usd(obj.total_amount)
    
    class Media:
        js = ('admin/inline_totals.js',)


@admin.register(SaleItem)
class SaleItemAdmin(admin.ModelAdmin):
    list_display = ('sale__id', 'product__name', 'show_quantity', 'sale_price_converted', 'total_cost', 'sale_date')
    list_display_links = ('sale__id', 'product__name')
    search_fields = ('sale__id', 'product__name')
    list_filter = (('sale__sale_date', DateRangeFilterBuilder()),  'product__brand__name')
    actions = (export_sale_items_to_excel,)

    @admin.display(description='Дата продажи')
    def sale_date(self, obj):
        return obj.sale.sale_date

    @admin.display(description='Количество')
    def show_quantity(self, obj):
        return f"{obj.quantity} шт."

    @admin.display(description='Цена продажи')
    def sale_price_converted(self, obj):
        return convert_from_usd(obj.sale_price)

    @admin.display(description='Итоговая стоимость')
    def total_cost(self, obj):
        return convert_from_usd(obj.total_cost)


@admin.register(Client)
class ClientAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'phone_number', 'balance', 'address')
    search_fields = ('full_name', 'phone_number')


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ('client__full_name', 'amount', 'payment_date')
    search_fields = ('client__full_name',)
    list_filter = ('payment_date',)
    autocomplete_fields = ('client',)

    def has_change_permission(self, request, obj=None):
        if obj:
            return False
        return super().has_change_permission(request, obj)