from django.contrib import admin
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.db.models import Sum, F, Q, DecimalField, ExpressionWrapper
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import now
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Warehouse, Country, Brand, Product, Arrival, ArrivalProduct, CurrencyRate
from .utils import convert_from_usd
from .filters import SoldQuantityFilter, SalePeriodFilter


@admin.register(ArrivalProduct)
class ArrivalProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'article_number', 'arrival__date', 'show_quantity', 'cost_price_converted', 
                    'total_cost_converted', 'brand__name', 'suits_for')
    list_display_links = ('name', 'article_number')
    search_fields = ('name', 'article_number')
    list_filter = ('arrival__date', 'brand__name')
    autocomplete_fields = ('arrival', 'brand')

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return (
                "name",
                "article_number",
                "brand",
                "arrival",
            )
        return () 

    def cost_price_converted(self, obj):
        return convert_from_usd(obj.cost_price)
    cost_price_converted.short_description = "–°–µ–±–µ—Å—Ç–æ–∏–º–æ—Å—Ç—å"

    def total_cost_converted(self, obj):
        return convert_from_usd(obj.total_cost)
    total_cost_converted.short_description = "–û–±—â–∞—è —Å—Ç–æ–∏–º–æ—Å—Ç—å"

    @admin.display(description='–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ')
    def show_quantity(self, obj):
        return f"{obj.quantity} —à—Ç."


class ArrivalProductInline(admin.TabularInline):
    model = ArrivalProduct
    extra = 5
    
    fields = (
        'product',
        'name',
        'article_number',
        'quantity',
        'cost_price',
        'row_total',
        'brand',
        'suits_for',
    )
    readonly_fields = ('row_total',)
    autocomplete_fields = ('product', 'brand',)
    can_delete = False

    def row_total(self, obj):
        return "‚Äî"
    row_total.short_description = "–ò—Ç–æ–≥–æ"

    class Media:
        js = ("admin/js/inline_row_numbers.js",)
        css = {
            "all": ("admin/css/inline_row_numbers.css",)
        }
    
    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)

        class ReadOnlyFormSet(formset):
            def __init__(self, *args, **inner_kwargs):
                super().__init__(*args, **inner_kwargs)

                for form in self.forms:
                    if form.instance.pk:
                        for field in ["name", "article_number", "brand", "arrival", 'row_total']:
                            if field in form.fields:
                                form.fields[field].disabled = True

        return ReadOnlyFormSet


@admin.register(Arrival)
class ArrivalAdmin(admin.ModelAdmin):
    list_display = ('id', 'date', 'warehouse__name', 'country_of_origin', 'total_amount_converted', 'comment')
    list_display_links = ('id', 'date')
    list_filter = ('warehouse__name', 'date')
    search_fields = ('id', 'warehouse__name', 'date')
    inlines = [ArrivalProductInline]

    def total_amount_converted(self, obj):
        return convert_from_usd(obj.total_amount)
    total_amount_converted.short_description = "–û–±—â–∞—è —Å—É–º–º–∞"

    class Media:
        js = (
            'admin/js/arrival_totals.js', 
            "admin/js/inline_keyboard_nav.js",
            "admin/js/arrival_product_autofill.js",
        )


@admin.action(description='–≠–∫—Å–ø–æ—Ä—Ç –æ—Å—Ç–∞—Ç–∫–æ–≤ —Å–∫–ª–∞–¥–∞ –≤ Excel')
def export_warehouse_stock_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "–û—Å—Ç–∞—Ç–∫–∏ —Å–∫–ª–∞–¥–∞"

    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    left = Alignment(horizontal='left')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ================= FILTER: ONLY quantity > 0 =================
    queryset = queryset.filter(quantity__gt=0).select_related(
        'warehouse', 'brand', 'country_of_origin'
    )

    # ================= WAREHOUSE INFO =================
    warehouse_ids = request.GET.get('warehouse__name')
    print(warehouse_ids, 'edcedc', request.GET, )
    if warehouse_ids:
        warehouse_text = warehouse_ids
    else:
        warehouse_text = "–í—Å–µ —Å–∫–ª–∞–¥—ã"

    # ================= DATE OF GENERATION =================
    generated_at = now().strftime('%Y-%m-%d %H:%M')

    # ================= TOTALS =================
    total_positions = queryset.count()
    total_quantity = sum(p.quantity for p in queryset)

    total_cost_amount = queryset.aggregate(
        total=Sum(
            ExpressionWrapper(
                F('quantity') * F('cost_price'),
                output_field=DecimalField(max_digits=15, decimal_places=2)
            )
        )
    )['total'] or 0

    # ================= TOP INFO =================
    ws.merge_cells('A1:H1')
    ws['A1'] = f"–°–∫–ª–∞–¥: {warehouse_text}"
    ws['A1'].font = bold
    ws['A1'].alignment = left

    ws.merge_cells('A2:H2')
    ws['A2'] = f"–î–∞—Ç–∞ —Ñ–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–∏—è –æ—Ç—á—ë—Ç–∞: {generated_at}"
    ws['A2'].font = bold
    ws['A2'].alignment = left

    ws.merge_cells('A3:H3')
    ws['A3'] = f"–í—Å–µ–≥–æ –ø–æ–∑–∏—Ü–∏–π: {total_positions} | –û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ: {total_quantity}"
    ws['A3'].font = bold
    ws['A3'].alignment = left

    ws.merge_cells('A4:H4')
    ws['A4'] = (
        f"–û–±—â–∞—è —Å–µ–±–µ—Å—Ç–æ–∏–º–æ—Å—Ç—å: {total_cost_amount:.2f} | "
    )
    ws['A4'].font = bold
    ws['A4'].alignment = left

    # ================= HEADERS =================
    headers = [
        "–¢–æ–≤–∞—Ä",
        "–ê—Ä—Ç–∏–∫—É–ª",
        "–°–∫–ª–∞–¥",
        "–ë—Ä–µ–Ω–¥",
        "–°—Ç—Ä–∞–Ω–∞",
        "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ",
        "–°–µ–±–µ—Å—Ç–æ–∏–º–æ—Å—Ç—å",
        "–¶–µ–Ω–∞ –ø—Ä–æ–¥–∞–∂–∏",
    ]

    ws.append([])
    ws.append(headers)

    header_row = ws.max_row

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border

    # ================= DATA =================
    for product in queryset:
        ws.append([
            product.name,
            product.article_number or '',
            product.warehouse.name,
            product.brand.name,
            product.country_of_origin.name,
            product.quantity,
            float(product.cost_price) if product.cost_price else '',
            float(product.selling_price) if product.selling_price else '',
        ])

    # ================= BODY STYLE =================
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = left if cell.column in [1, 3, 4, 5] else center

    # ================= AUTO WIDTH =================
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # ================= RESPONSE =================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="warehouse_stock.xlsx"'
    wb.save(response)

    return response


@admin.action(description="üí∞ –ü–æ–∫–∞–∑–∞—Ç—å –æ–±—â—É—é —Å–µ–±–µ—Å—Ç–æ–∏–º–æ—Å—Ç—å")
def show_total_cost_price(modeladmin, request, queryset):

    total_cost = queryset.aggregate(
        total=Sum(
            ExpressionWrapper(
                F('quantity') * F('cost_price'),
                output_field=DecimalField(max_digits=15, decimal_places=2)
            )
        )
    )['total'] or 0

    context = {
        'products': queryset.select_related('warehouse', 'brand'),
        'total_cost': convert_from_usd(total_cost),
    }

    return render(
        request,
        'admin/products/total_cost_price.html',
        context
    )


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'article_number', 'brand__name', 'country_of_origin', 'warehouse',  
                    'show_quantity', 'sold_quantity', 'cost_price_converted', 'suits_for')
    list_display_links = ('name', 'article_number')
    search_fields = ('name', 'article_number')
    list_filter = ('warehouse__name', 'country_of_origin__name', SoldQuantityFilter,
                   SalePeriodFilter, 'brand__name',)
    actions = (export_warehouse_stock_to_excel, show_total_cost_price)
    
    def cost_price_converted(self, obj):
        return convert_from_usd(obj.cost_price)
    cost_price_converted.short_description = "–°–µ–±–µ—Å—Ç–æ–∏–º–æ—Å—Ç—å"

    # def get_search_results(self, request, queryset, search_term):
    #     queryset, use_distinct = super().get_search_results(
    #         request, queryset, search_term
    #     )

    #     results = []
    #     for p in queryset:
    #         print(p, 'product in search', p.id )
    #         results.append({
    #             "id": p.id,
    #             "text": str(p),
    #             "article_number": p.article_number,
    #             "cost_price": str(p.cost_price or ""),
    #             "brand_id": p.brand_id,
    #         })
    #     print(results, 'search results')
    #     return results, use_distinct
    

    def get_queryset(self, request):
        qs = super().get_queryset(request).filter(quantity__gt=0)

        period = request.GET.get('period')
        now = timezone.now()

        sold_filter = Q()

        if period == 'today':
            sold_filter = Q(sale_items__sale__sale_date__gte=now.replace(
                hour=0, minute=0, second=0, microsecond=0
            ))

        elif period == 'week':
            start = now - timedelta(days=now.weekday())
            sold_filter = Q(sale_items__sale__sale_date__gte=start)

        elif period == 'month':
            sold_filter = Q(sale_items__sale__sale_date__gte=now.replace(day=1))

        elif period == 'year':
            sold_filter = Q(sale_items__sale__sale_date__gte=now.replace(month=1, day=1))

        return qs.annotate(
            active_sold_qty=Sum(
                'sale_items__quantity',
                filter=sold_filter,
                default=0
            )
        )

    @admin.display(description="–ü—Ä–æ–¥–∞–Ω–æ", ordering='active_sold_qty')
    def sold_quantity(self, obj):
        return obj.active_sold_qty or 0

    @admin.display(description="–ù–∞ —Å–∫–ª–∞–¥–µ")
    def show_quantity(self, obj):
        return f"{obj.quantity} —à—Ç."

    def get_ordering(self, request):
        if request.GET.get('sold') in ('most', 'least'):
            return ()
        return ('name',)


@admin.register(Warehouse)
class WarehouseAdmin(admin.ModelAdmin):
    list_display = ('name', 'total_products_type', 'total_quantity_of_goods')
    search_fields = ('name',)

    @admin.display(description='–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –≤–∏–¥–æ–≤ —Ç–æ–≤–∞—Ä–æ–≤')
    def total_products_type(self, obj):
        return obj.total_products_type
    
    @admin.display(description='–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ–≤–∞—Ä–æ–≤')
    def total_quantity_of_goods(self, obj):
        return obj.total_quantity_of_goods


@admin.register(Country)
class CountryAdmin(admin.ModelAdmin):
    list_display = ('name', 'flag_image_display')
    search_fields = ('name',)

    @admin.display(description='–§–ª–∞–≥')
    def flag_image_display(self, obj):
        if obj.flag_image:
            return mark_safe(f'<img src="{obj.flag_image.url}" width="100" height="60" />')
        return "-"


@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ('name', 'logo_display')
    search_fields = ('name',)

    @admin.display(description='–õ–æ–≥–æ—Ç–∏–ø')
    def logo_display(self, obj):
        if obj.logo:
            return mark_safe(f'<img src="{obj.logo.url}" width="100" height="70" />')
        return "-"
    

@admin.register(CurrencyRate)
class CurrencyRateAdmin(admin.ModelAdmin):
    list_display = ('currency_code', 'rate_to_usd', 'last_updated', 'selected')
    search_fields = ('currency_code',)
    list_filter = ('last_updated',)
    list_editable = ('selected',)